from __future__ import annotations

import csv
import json
import os
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import duckdb

from . import assertions
from .cleaning import dfloat, is_missing, run_all_cleaning
from .config import DEFAULT_DATA_DIR, DEFAULT_DUCKDB_PATH, SQL_DIR

STG_TABLES = [
    ("stg_vessel", "vessels.csv"),
    ("stg_voyage", "voyages.csv"),
    ("stg_freight_invoice", "freight_invoices.csv"),
    ("stg_port_cost", "port_costs.csv"),
    ("stg_bunker_stem", "bunker_stems.csv"),
    ("stg_laytime_statement", "laytime_statements.csv"),
    ("stg_worldscale_flat_rates", "worldscale_flat_rates.csv"),
    ("stg_open_positions", "open_positions.csv"),
]


@dataclass(frozen=True)
class PortKey:
    name: str
    country: str | None
    region: str | None
    is_virtual: bool


def connect_duckdb(db_path: Path) -> duckdb.DuckDBPyConnection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    if db_path.exists():
        db_path.unlink()
    return duckdb.connect(str(db_path))


def apply_mart_ddl(con: duckdb.DuckDBPyConnection) -> None:
    sql = (SQL_DIR / "duck_schema.sql").read_text(encoding="utf-8")
    con.execute(sql)


def load_staging_from_cleaned(con: duckdb.DuckDBPyConnection, cleaned_dir: Path) -> None:
    for table, fn in STG_TABLES:
        path = cleaned_dir / fn
        con.execute(
            f"CREATE OR REPLACE TABLE {table} AS SELECT * FROM read_csv_auto(?)",
            [str(path)],
        )


def load_broker_staging(con: duckdb.DuckDBPyConnection, cleaned_dir: Path) -> None:
    path = cleaned_dir / "broker_messages.jsonl"
    con.execute("DROP TABLE IF EXISTS stg_broker_messages")
    con.execute(
        """CREATE TABLE stg_broker_messages (
            line_no BIGINT, message_id VARCHAR, received_at VARCHAR,
            from_addr VARCHAR, subject VARCHAR, body VARCHAR
        )"""
    )
    rows: list[tuple] = []
    with path.open(encoding="utf-8") as f:
        for i, line in enumerate(f, start=1):
            line = line.strip()
            if not line:
                continue
            obj = json.loads(line)
            rows.append(
                (
                    i,
                    str(obj["message_id"]),
                    str(obj["received_at"]),
                    str(obj["from"]),
                    str(obj["subject"]),
                    str(obj["body"]),
                )
            )
    if rows:
        con.executemany(
            "INSERT INTO stg_broker_messages VALUES (?,?,?,?,?,?)",
            rows,
        )


def load_excel_staging_tables(con: duckdb.DuckDBPyConnection, data_dir: Path) -> None:
    from openpyxl import load_workbook

    dem_rows: list[tuple] = []
    bb_rows: list[tuple] = []
    xlsx = data_dir / "excel_tracker.xlsx"
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["Demurrage Claims"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        low = [v.lower() for v in vals]
        if "voy ref" in low and "claimed (usd)" in low:
            header_idx = i
            break
    assert header_idx is not None
    hdr = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col = {h.lower(): j for j, h in enumerate(hdr)}
    for j, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        def g(name: str):
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        dem_rows.append(
            (
                j,
                None if g("voy ref") is None else str(g("voy ref")).strip(),
                None if g("vessel") is None else str(g("vessel")).strip(),
                None if g("counterparty") is None else str(g("counterparty")).strip(),
                None if g("type") is None else str(g("type")).strip(),
                dfloat(g("claimed (usd)")),
                dfloat(g("agreed (usd)")),
                None if g("status") is None else str(g("status")).strip(),
                dfloat(g("days outstanding")),
            )
        )
    ws = wb["Bunker Budget"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        low = [v.lower() for v in vals]
        if "voyage id" in low and "budget mt" in low:
            header_idx = i
            break
    assert header_idx is not None
    hdr = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col = {h.lower(): j for j, h in enumerate(hdr)}
    for j, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue

        def g2(name: str):
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        bb_rows.append(
            (
                j,
                None if g2("voyage id") is None else str(g2("voyage id")).strip(),
                None if g2("vessel") is None else str(g2("vessel")).strip(),
                None if g2("grade") is None else str(g2("grade")).strip(),
                dfloat(g2("budget mt")),
                dfloat(g2("budget $/mt")),
                dfloat(g2("actual mt")),
                dfloat(g2("actual $/mt")),
                dfloat(g2("variance $")),
            )
        )
    wb.close()

    con.execute(
        """CREATE OR REPLACE TABLE stg_excel_demurrage_claims (
            row_no BIGINT, voy_ref VARCHAR, vessel VARCHAR, counterparty VARCHAR,
            claim_type VARCHAR, claimed_usd DOUBLE, agreed_usd DOUBLE,
            status VARCHAR, days_outstanding DOUBLE
        )"""
    )
    if dem_rows:
        con.executemany(
            "INSERT INTO stg_excel_demurrage_claims VALUES (?,?,?,?,?,?,?,?,?)",
            dem_rows,
        )

    con.execute(
        """CREATE OR REPLACE TABLE stg_excel_bunker_budget (
            row_no BIGINT, voyage_id VARCHAR, vessel VARCHAR, grade VARCHAR,
            budget_mt DOUBLE, budget_usd_per_mt DOUBLE, actual_mt DOUBLE,
            actual_usd_per_mt DOUBLE, variance_usd DOUBLE
        )"""
    )
    if bb_rows:
        con.executemany(
            "INSERT INTO stg_excel_bunker_budget VALUES (?,?,?,?,?,?,?,?,?)",
            bb_rows,
        )


def _table_dicts(con: duckdb.DuckDBPyConnection, sql: str) -> list[dict[str, Any]]:
    res = con.execute(sql)
    desc = res.description
    cols = [c[0] for c in desc] if desc else []
    return [dict(zip(cols, row)) for row in res.fetchall()]


def build_port_registry(con: duckdb.DuckDBPyConnection) -> dict[PortKey, int]:
    keys: list[PortKey] = []
    voy = _table_dicts(con, "SELECT * FROM stg_voyage")
    for r in voy:
        keys.append(
            PortKey(r["load_port"].strip(), r["load_country"].strip(), r["load_region"].strip(), False)
        )
        keys.append(
            PortKey(
                r["discharge_port"].strip(),
                r["discharge_country"].strip(),
                r["discharge_region"].strip(),
                r["discharge_port"].strip().upper() == "STS",
            )
        )
        keys.append(PortKey(r["ballast_origin"].strip(), None, None, False))
    for (bp,) in con.execute("SELECT DISTINCT bunker_port FROM stg_bunker_stem").fetchall():
        keys.append(PortKey(str(bp).strip(), None, None, False))
    for (port,) in con.execute(
        "SELECT DISTINCT port FROM stg_port_cost WHERE port IS NOT NULL AND trim(port) != ''"
    ).fetchall():
        keys.append(PortKey(str(port).strip(), None, None, False))
    for lp, dp in con.execute(
        "SELECT DISTINCT load_port, disch_port FROM stg_open_positions"
    ).fetchall():
        if lp and str(lp).strip():
            keys.append(PortKey(str(lp).strip(), None, None, False))
        if dp and str(dp).strip():
            keys.append(PortKey(str(dp).strip(), None, None, False))

    uniq: dict[PortKey, int] = {}
    pid = 1
    for k in sorted(set(keys), key=lambda x: (x.is_virtual, x.name, x.country or "", x.region or "")):
        uniq[k] = pid
        pid += 1
    return uniq


def insert_ports(con: duckdb.DuckDBPyConnection, reg: dict[PortKey, int]) -> None:
    for pk, pid in sorted(reg.items(), key=lambda x: x[1]):
        con.execute(
            """INSERT INTO Port (port_id, port_name, country, region, is_virtual_port)
               VALUES (?,?,?,?,?)""",
            [pid, pk.name, pk.country, pk.region, pk.is_virtual],
        )


def insert_vessel_charterer_cargo_date(con: duckdb.DuckDBPyConnection) -> None:
    for r in _table_dicts(con, "SELECT * FROM stg_vessel"):
        con.execute(
            """INSERT INTO Vessel VALUES (?,?,?,?,?,?,?)""",
            [
                r["imo_number"],
                r["vessel_name"],
                r["vessel_type"],
                float(r["dwt_mt"]),
                int(r["build_year"]),
                r["flag_state"],
                r["scrubber_fitted"] == "Y",
            ],
        )

    for i, (name,) in enumerate(
        con.execute("SELECT DISTINCT charterer FROM stg_voyage ORDER BY charterer").fetchall(), start=1
    ):
        con.execute(
            "INSERT INTO Charterer (charterer_id, charterer_name) VALUES (?,?)",
            [i, str(name).strip()],
        )

    for i, (g,) in enumerate(
        con.execute("SELECT DISTINCT cargo_grade FROM stg_voyage ORDER BY cargo_grade").fetchall(), start=1
    ):
        con.execute("INSERT INTO Cargo (cargo_id, cargo_grade) VALUES (?,?)", [i, str(g).strip()])

    dates: set[str] = set()
    for r in _table_dicts(con, "SELECT * FROM stg_voyage"):
        for c in (
            "cp_date",
            "laycan_start",
            "laycan_end",
            "actual_load_date",
            "actual_discharge_date",
        ):
            if not is_missing(r.get(c)):
                dates.add(str(r[c])[:10])
        try:
            ld = datetime.strptime(str(r["actual_load_date"])[:10], "%Y-%m-%d").date()
            bd = int(float(r["ballast_days"]))
            dates.add((ld - timedelta(days=bd)).isoformat())
        except (ValueError, TypeError, KeyError):
            pass
    for r in _table_dicts(con, "SELECT invoice_date, payment_date FROM stg_freight_invoice"):
        for k in ("invoice_date", "payment_date"):
            if not is_missing(r.get(k)):
                dates.add(str(r[k])[:10])
    for (s,) in con.execute("SELECT stem_date FROM stg_bunker_stem").fetchall():
        if s:
            dates.add(str(s)[:10])
    for (s,) in con.execute("SELECT call_date FROM stg_port_cost").fetchall():
        if s:
            dates.add(str(s)[:10])
    for r in _table_dicts(con, "SELECT nor_tendered, commencement FROM stg_laytime_statement"):
        for k in ("nor_tendered", "commencement"):
            if not is_missing(r.get(k)):
                dates.add(str(r[k])[:10])

    for ds in sorted(dates):
        d = datetime.strptime(ds, "%Y-%m-%d").date()
        q = (d.month - 1) // 3 + 1
        con.execute(
            "INSERT INTO DateDim VALUES (?,?,?,?,?)",
            [ds, d.year, q, d.month, d.day],
        )


def port_id_lookup(reg: dict[PortKey, int], name: str, country: str | None, region: str | None) -> int:
    virt = name.strip().upper() == "STS" and (country or "").strip().upper() == "STS"
    pk = PortKey(name.strip(), country, region, virt)
    if pk not in reg:
        pk = PortKey(name.strip(), country, region, False)
    if pk not in reg:
        raise KeyError(f"Port not in registry: {name!r} {country!r} {region!r}")
    return reg[pk]


def load_est_tce_map(cleaned_dir: Path) -> dict[str, float]:
    out: dict[str, float] = {}
    path = cleaned_dir / "open_positions.csv"
    with path.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            vr = (r.get("voy_ref") or "").strip()
            if not vr:
                continue
            est = dfloat(r.get("est_tce_usd_day"))
            if est is not None:
                out[vr] = float(est)
    return out


def run_pipeline(data_dir: Path | None = None, db_path: Path | None = None) -> Path:
    data_dir = data_dir or Path(os.environ.get("REEDERI_DATA_DIR", str(DEFAULT_DATA_DIR)))
    db_path = db_path or Path(os.environ.get("REEDERI_DB_PATH", str(DEFAULT_DUCKDB_PATH)))
    cleaned_dir = data_dir / "output" / "cleaned"
    paths = run_all_cleaning(data_dir, cleaned_dir)

    assertions.step_clean_voyages_no_null_discharge(paths["voyages"])
    assertions.step_clean_laytime_no_null_amount(paths["laytime"])
    assertions.step_clean_port_no_null_port(paths["port_costs"])
    assertions.step_clean_open_positions_voy_ref_max_missing(paths["open_positions"], max_missing=3)

    con = connect_duckdb(db_path)
    try:
        apply_mart_ddl(con)
        load_staging_from_cleaned(con, cleaned_dir)
        load_broker_staging(con, cleaned_dir)
        load_excel_staging_tables(con, data_dir)

        assertions.step_assert_staging_counts_duck(con)
        assertions.step_assert_one_invoice_per_voyage(con)
        assertions.step_assert_one_laytime_per_voyage(con)

        reg = build_port_registry(con)
        insert_ports(con, reg)
        insert_vessel_charterer_cargo_date(con)

        charterer_by = {
            a: b
            for a, b in con.execute("SELECT charterer_name, charterer_id FROM Charterer").fetchall()
        }
        cargo_by = {a: b for a, b in con.execute("SELECT cargo_grade, cargo_id FROM Cargo").fetchall()}
        freight_by = {
            str(a).strip(): float(b)
            for a, b in con.execute("SELECT voyage_id, gross_freight_usd FROM stg_freight_invoice").fetchall()
        }
        bunker_by = {
            str(a).strip(): float(b)
            for a, b in con.execute(
                "SELECT voyage_id, SUM(total_cost_usd) FROM stg_bunker_stem GROUP BY voyage_id"
            ).fetchall()
        }
        lay_by = {
            str(r["voyage_id"]).strip(): r
            for r in _table_dicts(con, "SELECT * FROM stg_laytime_statement")
        }

        pc_rows = con.execute(
            """
            SELECT voyage_id, lower(port_type) AS pt,
              SUM(agency_fee + pilotage + towage + port_dues + mooring) AS fees,
              SUM(canal_transit) AS canal
            FROM stg_port_cost
            GROUP BY 1, 2
            """
        ).fetchall()
        port_split: dict[str, dict[str, float]] = {}
        for vid, pt, fees, canal in pc_rows:
            port_split.setdefault(str(vid).strip(), {})[str(pt).strip()] = {
                "fees": float(fees or 0),
                "canal": float(canal or 0),
            }

        est_map = load_est_tce_map(cleaned_dir)

        voy_list = _table_dicts(con, "SELECT * FROM stg_voyage")
        leg_id = 1
        for r in voy_list:
            vid = str(r["voyage_id"]).strip()
            imo = str(r["imo_number"]).strip()
            ch = charterer_by[str(r["charterer"]).strip()]
            cg = cargo_by[str(r["cargo_grade"]).strip()]
            sts = str(r["sts_transfer"]).strip() == "Y"
            disputed = str(lay_by[vid]["disputed"]).strip() == "Y"
            laden_d = float(r["laden_days"])
            ballast_d = float(r["ballast_days"])
            total_d = laden_d + ballast_d
            gross = float(freight_by[vid])
            bunk_total = float(bunker_by.get(vid) or 0.0)
            sh = port_split.get(vid, {})
            load_fees = sh.get("load", {}).get("fees", 0.0)
            load_canal = sh.get("load", {}).get("canal", 0.0)
            disc_fees = sh.get("discharge", {}).get("fees", 0.0)
            disc_canal = sh.get("discharge", {}).get("canal", 0.0)

            load_pid = port_id_lookup(reg, r["load_port"], r["load_country"], r["load_region"])
            disc_pid = port_id_lookup(
                reg, r["discharge_port"], r["discharge_country"], r["discharge_region"]
            )
            ballast_pid = port_id_lookup(reg, r["ballast_origin"], None, None)

            load_date = str(r["actual_load_date"])[:10]
            disc_date = str(r["actual_discharge_date"])[:10]
            ballast_start = (
                datetime.strptime(load_date, "%Y-%m-%d") - timedelta(days=int(ballast_d))
            ).date()

            lay = lay_by[vid]

            def _nor_date(val: Any) -> date:
                s = str(val).strip().replace("T", " ")
                if len(s) >= 19 and s[16] == ":":
                    return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S").date()
                return datetime.strptime(s[:16], "%Y-%m-%d %H:%M").date()

            nor_d = _nor_date(lay["nor_tendered"])
            load_d = datetime.strptime(load_date, "%Y-%m-%d").date()
            assign_ballast = nor_d == load_d
            nh = float(lay["net_hours"])
            mag = abs(float(lay["amount_usd"]))
            signed_dem = mag if nh < 0 else -mag

            dem_ballast = signed_dem if assign_ballast else None
            dem_laden = signed_dem if not assign_ballast else None

            share_l = laden_d / total_d if total_d else 0.0
            share_b = ballast_d / total_d if total_d else 0.0

            est = est_map.get(vid)

            def tce_row(alloc_f: float, dem: float | None, bunk: float, pport: float, pcanal: float, days: float) -> float:
                d = 0.0 if dem is None else float(dem)
                return (alloc_f - d - bunk - pport - pcanal) / days

            alloc_l = gross * share_l
            bunk_l = bunk_total * share_l
            tce_l = tce_row(alloc_l, dem_laden, bunk_l, disc_fees, disc_canal, laden_d)

            con.execute(
                """INSERT INTO Voyage_Leg (
                  leg_id, voyage_id, imo_number, charterer_id, cargo_id,
                  origin_port_id, destination_port_id, start_date, end_date,
                  leg_type, sts_transfer, disputed, leg_days,
                  allocated_freight_usd, bunker_cost_usd, port_cost_usd, canal_transit_usd,
                  demurrage_cost_usd, tce, est_tce
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [
                    leg_id,
                    f"{vid}-L",
                    imo,
                    ch,
                    cg,
                    load_pid,
                    disc_pid,
                    load_date,
                    disc_date,
                    "Laden",
                    sts,
                    disputed,
                    laden_d,
                    alloc_l,
                    bunk_l,
                    disc_fees,
                    disc_canal,
                    dem_laden,
                    tce_l,
                    est,
                ],
            )
            leg_id += 1

            alloc_b = gross * share_b
            bunk_b = bunk_total * share_b
            tce_b = tce_row(alloc_b, dem_ballast, bunk_b, load_fees, load_canal, ballast_d)

            con.execute(
                """INSERT INTO Voyage_Leg (
                  leg_id, voyage_id, imo_number, charterer_id, cargo_id,
                  origin_port_id, destination_port_id, start_date, end_date,
                  leg_type, sts_transfer, disputed, leg_days,
                  allocated_freight_usd, bunker_cost_usd, port_cost_usd, canal_transit_usd,
                  demurrage_cost_usd, tce, est_tce
                ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [
                    leg_id,
                    f"{vid}-B",
                    imo,
                    ch,
                    cg,
                    ballast_pid,
                    load_pid,
                    ballast_start,
                    load_date,
                    "Ballast",
                    sts,
                    disputed,
                    ballast_d,
                    alloc_b,
                    bunk_b,
                    load_fees,
                    load_canal,
                    dem_ballast,
                    tce_b,
                    est,
                ],
            )
            leg_id += 1

        assertions.step_assert_mart_dims(con)
        assertions.step_assert_mart_voyage_legs(con)
        assertions.step_assert_allocated_freight_matches_invoice(con)
        assertions.step_assert_bunker_split_matches_stem_total(con)
        assertions.step_assert_port_canal_matches_staging(con)

        con.commit()
    finally:
        con.close()

    return db_path


def main() -> None:
    data_dir = Path(os.environ.get("REEDERI_DATA_DIR", str(DEFAULT_DATA_DIR)))
    out = run_pipeline()
    cleaned = (data_dir / "output" / "cleaned").resolve()
    print(f"ETL complete: {out.resolve()}")
    print(f"Cleaned datasets: {cleaned}")


if __name__ == "__main__":
    main()
