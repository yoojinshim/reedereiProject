from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .vessel_normalize import normalize_vessel


def is_missing(v: Any) -> bool:
    if v is None:
        return True
    s = str(v).strip().lower()
    return s in {"", "na", "n/a", "null", "none"}


def dfloat(x: Any) -> float | None:
    if is_missing(x):
        return None
    try:
        return float(str(x).replace(",", "").strip())
    except ValueError:
        return None


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)


def clean_voyages(data_dir: Path, out_path: Path) -> list[dict]:
    rows_out: list[dict] = []
    with (data_dir / "voyages.csv").open(newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            row = dict(r)
            if is_missing(row.get("discharge_port")):
                row["discharge_port"] = "STS"
            if is_missing(row.get("discharge_country")):
                row["discharge_country"] = "STS"
            if is_missing(row.get("discharge_region")):
                row["discharge_region"] = "STS"
            rows_out.append(row)
    write_csv(out_path, list(rows_out[0].keys()), rows_out)
    return rows_out


def load_demurrage_claims(data_dir: Path) -> dict[str, dict]:
    """voy_ref -> {claimed_usd, agreed_usd, status}."""
    wb = load_workbook(data_dir / "excel_tracker.xlsx", data_only=True, read_only=True)
    ws = wb["Demurrage Claims"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        low = [v.lower() for v in vals]
        if "voy ref" in low and "claimed (usd)" in low:
            header_idx = i
            break
    if header_idx is None:
        wb.close()
        raise RuntimeError("Demurrage Claims header not found")
    hdr = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    col = {h.lower(): j for j, h in enumerate(hdr)}
    out: dict[str, dict] = {}
    for row in rows[header_idx + 1 :]:
        if row is None:
            continue
        def g(name: str):
            idx = col.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        vr = g("voy ref")
        if vr is None or str(vr).strip() == "":
            continue
        vid = str(vr).strip()
        out[vid] = {
            "claimed_usd": dfloat(g("claimed (usd)")),
            "agreed_usd": dfloat(g("agreed (usd)")),
            "status": None if g("status") is None else str(g("status")).strip(),
        }
    wb.close()
    return out


def impute_laytime_amount_usd(
    claimed: float | None, agreed: float | None, status: str | None,
) -> float | None:
    st = (status or "").strip()
    if agreed is not None:
        return float(agreed)
    if st.lower() == "agreed" and claimed is not None:
        return float(claimed)
    if claimed is not None:
        return float(claimed) * 0.75
    return None


def clean_laytime(data_dir: Path, out_path: Path) -> list[dict]:
    claims = load_demurrage_claims(data_dir)
    rows_out: list[dict] = []
    with (data_dir / "laytime_statements.csv").open(newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            row = dict(r)
            amt = dfloat(row.get("amount_usd"))
            if amt is None:
                vid = row["voyage_id"].strip()
                c = claims.get(vid)
                if c is None:
                    raise ValueError(f"Laytime amount null but no Demurrage Claims row for {vid}")
                imputed = impute_laytime_amount_usd(
                    c["claimed_usd"], c["agreed_usd"], c["status"],
                )
                if imputed is None:
                    raise ValueError(f"Could not impute amount_usd for voyage {vid}")
                row["amount_usd"] = f"{imputed:.2f}"
            rows_out.append(row)
    write_csv(out_path, list(rows_out[0].keys()), rows_out)
    return rows_out


def clean_port_costs(data_dir: Path, out_path: Path) -> list[dict]:
    rows_out: list[dict] = []
    with (data_dir / "port_costs.csv").open(newline="", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            row = dict(r)
            if is_missing(row.get("port")):
                row["port"] = "STS"
            rows_out.append(row)
    write_csv(out_path, list(rows_out[0].keys()), rows_out)
    return rows_out


def load_raw_open_positions(data_dir: Path) -> tuple[list[str], list[tuple]]:
    wb = load_workbook(data_dir / "excel_tracker.xlsx", data_only=True, read_only=True)
    ws = wb["Open Positions"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        vals = [str(c).strip() if c is not None else "" for c in row]
        if "voy ref (erp)" in [v.lower() for v in vals]:
            header_idx = i
            break
    if header_idx is None:
        wb.close()
        raise RuntimeError("Open Positions header not found")
    hdr = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
    data_rows = []
    for j, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        data_rows.append((j, row))
    wb.close()
    return hdr, data_rows


def clean_open_positions(
    data_dir: Path,
    voyages: list[dict],
    invoices: list[dict],
    vessels: list[dict],
    out_path: Path,
) -> list[dict]:
    hdr, data_rows = load_raw_open_positions(data_dir)
    col = {h.lower(): j for j, h in enumerate(hdr)}

    vessel_to_imo = {v["vessel_name"].strip(): v["imo_number"].strip() for v in vessels}

    def g(row: tuple, name: str):
        idx = col.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    inv_ws = {r["voyage_id"].strip(): dfloat(r["worldscale_points"]) for r in invoices}

    def voyage_match_score(v: dict, charterer: str, vessel_norm: str, lp: str, dp: str, gr: str | None, qty: float | None) -> int:
        sc = 0
        if v["charterer"].strip() == charterer:
            sc += 25
        if v["vessel_name"].strip() == vessel_norm:
            sc += 25
        if v["load_port"].strip() == lp:
            sc += 15
        dpv = (v.get("discharge_port") or "").strip()
        if (dp or "").strip() == dpv:
            sc += 15
        if gr and v["cargo_grade"].strip() == gr.strip():
            sc += 10
        if qty is not None and dfloat(v.get("cargo_qty_mt")) is not None:
            cq = float(v["cargo_qty_mt"])
            if cq > 0 and abs(cq - qty) / cq <= 0.02:
                sc += 10
        return sc

    out_rows: list[dict] = []
    for row_no, row in data_rows:
        vessel_raw = g(row, "vessel")
        vessel_norm = normalize_vessel(vessel_raw) if vessel_raw else None
        imo_number = vessel_to_imo.get(vessel_norm) if vessel_norm else None
        voy_ref = g(row, "voy ref (erp)")
        voy_s = str(voy_ref).strip() if voy_ref is not None and str(voy_ref).strip() else None
        charterer = str(g(row, "charterer")).strip() if g(row, "charterer") else ""
        lp = str(g(row, "load port")).strip() if g(row, "load port") else ""
        raw_dp = g(row, "disch port")
        if is_missing(raw_dp):
            dp = "STS"
        else:
            dp = str(raw_dp).strip()
        grade = str(g(row, "grade")).strip() if g(row, "grade") else None
        if grade == "":
            grade = None
        qty = dfloat(g(row, "qty (mt)"))
        ws_rate = dfloat(g(row, "ws rate"))

        if not voy_s and vessel_norm and charterer:
            pool = [v for v in voyages if v["vessel_name"].strip() == vessel_norm]
            if not pool:
                pool = []
            scored: list[tuple[int, dict]] = []
            for v in (pool if pool else []):
                sc = voyage_match_score(v, charterer, vessel_norm, lp, dp, grade, qty)
                scored.append((sc, v))
            scored.sort(key=lambda x: -x[0])
            if scored:
                top_sc, top_v = scored[0]
                second_sc = scored[1][0] if len(scored) > 1 else -1
                unique_best = top_sc > second_sc
                if top_sc >= 50 or (unique_best and top_sc >= 40):
                    voy_s = top_v["voyage_id"].strip()

        if voy_s:
            vmatch = next((v for v in voyages if v["voyage_id"].strip() == voy_s), None)
            if vmatch:
                if not grade:
                    grade = vmatch["cargo_grade"].strip()
                if qty is None and dfloat(vmatch.get("cargo_qty_mt")) is not None:
                    qty = float(vmatch["cargo_qty_mt"])
                if ws_rate is None and voy_s in inv_ws and inv_ws[voy_s] is not None:
                    ws_rate = inv_ws[voy_s]

        out_rows.append(
            {
                "row_no": row_no,
                "vessel_raw": None if vessel_raw is None else str(vessel_raw).strip(),
                "vessel": vessel_norm,
                "imo_number": imo_number,
                "voy_ref": voy_s,
                "charterer": charterer or None,
                "load_port": lp or None,
                "disch_port": dp or None,
                "grade": grade,
                "qty_mt": qty,
                "ws_rate": ws_rate,
                "est_tce_usd_day": dfloat(g(row, "est. tce ($/d)")),
                "status": None if g(row, "status") is None else str(g(row, "status")).strip(),
                "notes": None if g(row, "notes / internal ref") is None else str(g(row, "notes / internal ref")).strip(),
            }
        )

    fields = list(out_rows[0].keys()) if out_rows else [
        "row_no",
        "vessel_raw",
        "vessel",
        "imo_number",
        "voy_ref",
        "charterer",
        "load_port",
        "disch_port",
        "grade",
        "qty_mt",
        "ws_rate",
        "est_tce_usd_day",
        "status",
        "notes",
    ]
    write_csv(out_path, fields, out_rows)
    return out_rows


def copy_unchanged_csv(src: Path, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    dst.write_bytes(src.read_bytes())


def copy_broker_jsonl(data_dir: Path, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text((data_dir / "broker_messages.jsonl").read_text(encoding="utf-8"), encoding="utf-8")


def run_all_cleaning(data_dir: Path, cleaned_dir: Path) -> dict[str, Path]:
    cleaned_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    voyages = clean_voyages(data_dir, cleaned_dir / "voyages.csv")
    paths["voyages"] = cleaned_dir / "voyages.csv"

    clean_laytime(data_dir, cleaned_dir / "laytime_statements.csv")
    paths["laytime"] = cleaned_dir / "laytime_statements.csv"

    clean_port_costs(data_dir, cleaned_dir / "port_costs.csv")
    paths["port_costs"] = cleaned_dir / "port_costs.csv"

    with (data_dir / "freight_invoices.csv").open(newline="", encoding="utf-8-sig") as f:
        invoices = list(csv.DictReader(f))
    with (data_dir / "vessels.csv").open(newline="", encoding="utf-8-sig") as f:
        vessels = list(csv.DictReader(f))
    
    clean_open_positions(data_dir, voyages, invoices, vessels, cleaned_dir / "open_positions.csv")
    paths["open_positions"] = cleaned_dir / "open_positions.csv"

    for name, fn in [
        ("vessels", "vessels.csv"),
        ("freight_invoices", "freight_invoices.csv"),
        ("bunker_stems", "bunker_stems.csv"),
        ("worldscale_flat_rates", "worldscale_flat_rates.csv"),
    ]:
        copy_unchanged_csv(data_dir / fn, cleaned_dir / fn)
        paths[name] = cleaned_dir / fn

    copy_broker_jsonl(data_dir, cleaned_dir / "broker_messages.jsonl")
    paths["broker_messages"] = cleaned_dir / "broker_messages.jsonl"

    excel_out = cleaned_dir / "excel_tracker_paths.json"
    excel_out.write_text(
        json.dumps(
            {
                "source": str((data_dir / "excel_tracker.xlsx").resolve()),
                "note": "Raw workbook unchanged; cleaned Open Positions exported as open_positions.csv",
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    paths["excel_meta"] = excel_out

    return paths
